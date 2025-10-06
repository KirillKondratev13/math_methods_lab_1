import time
from collections import Counter
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ---------------------------- Labels and constants ----------------------------

# Human-readable labels used for logging and selection
LABEL_MEAN = "Среднее значение"
LABEL_MEDIAN = "Медиана"
LABEL_MODE = "Мода"
LABEL_CONST = "Константа"
LABEL_FFILL = "Прямое заполнение (ffill)"
LABEL_BFILL = "Обратное заполнение (bfill)"
LABEL_INTERP = "Интерполяция"
LABEL_DROP_ROWS = "Удаление строк"
LABEL_UNCHANGED = "Без изменений"

# Allowed labels set (for validation or mapping if needed)
ALLOWED_METHOD_LABELS = {
    LABEL_MEAN,
    LABEL_MEDIAN,
    LABEL_MODE,
    LABEL_CONST,
    LABEL_FFILL,
    LABEL_BFILL,
    LABEL_INTERP,
    LABEL_DROP_ROWS,
    LABEL_UNCHANGED,
}


# ---------------------------- Helper utilities ----------------------------

def get_dtype_category(s: pd.Series) -> str:
    """
    Определяет категорию типа данных серии.
    Возвращает одно из: 'numeric', 'boolean', 'datetime', 'string'.
    Совместимо с разными версиями pandas (без is_boolean_dtype и is_datetime64tz_dtype).
    """
    import pandas.api.types as ptypes

    dt = s.dtype

    # Datetime (including timezone-aware) detection compatible across pandas versions
    is_dt = False
    try:
        is_dt = bool(ptypes.is_datetime64_any_dtype(dt))
    except Exception:
        try:
            is_dt = bool(ptypes.is_datetime64_any_dtype(s))
        except Exception:
            is_dt = False
    if not is_dt:
        try:
            # Works across many pandas versions
            from pandas.core.dtypes.dtypes import DatetimeTZDtype  # type: ignore
            if isinstance(dt, DatetimeTZDtype):
                is_dt = True
        except Exception:
            pass
    if is_dt:
        return "datetime"

    # Boolean
    if ptypes.is_bool_dtype(dt) or ptypes.is_bool_dtype(s):
        return "boolean"

    # Numeric
    if ptypes.is_numeric_dtype(dt) or ptypes.is_numeric_dtype(s):
        return "numeric"

    # String-like (object, string extension, category)
    try:
        is_cat = bool(ptypes.is_categorical_dtype(dt))
    except Exception:
        is_cat = False
    if ptypes.is_string_dtype(dt) or ptypes.is_object_dtype(dt) or is_cat:
        return "string"

    # Fallback
    return "string"


def compute_mode_value(s: pd.Series) -> Tuple[Any, int]:
    """
    Вычисляет моду ряда.
    Возвращает (значение_моды, количество_равновероятных_мод).
    Если мод нет (все значения NaN), возвращает (None, 0).
    Для нескольких мод выбирается первая после сортировки по значению (строковому представлению).
    """
    vc = s.dropna().value_counts()
    if vc.empty:
        return None, 0
    top = vc.max()
    candidates = vc[vc == top].index.tolist()
    tie_count = len(candidates)
    # Стабильно сортируем по строковому виду
    candidates_sorted = sorted(candidates, key=lambda x: str(x))
    return candidates_sorted[0], tie_count


def apply_method_once(
    s: pd.Series,
    method_label: str,
    params: Optional[Dict[str, Any]] = None,
) -> Tuple[pd.Series, Dict[str, Any], bool]:
    """
    Применяет одну стратегию заполнения к серии без fallback'ов.
    Возвращает (новая_серия, параметры_использованные, применимо_ли).
    Параметры метода фиксируются для лога.
    """
    params = dict(params or {})
    dtype_cat = get_dtype_category(s)

    if method_label == LABEL_MEAN:
        # Applicable only to numeric
        if dtype_cat != "numeric":
            return s.copy(), params, False
        val = s.mean(skipna=True)
        if pd.isna(val):
            return s.copy(), params, False
        params["constant_value"] = val
        return s.fillna(value=val), params, True

    if method_label == LABEL_MEDIAN:
        # Applicable only to numeric
        if dtype_cat != "numeric":
            return s.copy(), params, False
        val = s.median(skipna=True)
        if pd.isna(val):
            return s.copy(), params, False
        params["constant_value"] = val
        return s.fillna(value=val), params, True

    if method_label == LABEL_MODE:
        # Applicable to any dtype if there is at least one non-null value
        val, tie_count = compute_mode_value(s)
        if val is None:
            return s.copy(), params, False
        params["tie_count"] = tie_count
        params["constant_value"] = val
        return s.fillna(value=val), params, True

    if method_label == LABEL_CONST:
        # Applicable if a constant is provided
        if "constant_value" not in params:
            return s.copy(), params, False
        return s.fillna(value=params["constant_value"]), params, True

    if method_label == LABEL_FFILL:
        limit = params.get("limit", None)
        # Forward fill generally applicable; may not change anything if no forward values available
        filled = s.fillna(method="ffill", limit=limit)
        params["limit"] = limit
        return filled, params, True

    if method_label == LABEL_BFILL:
        limit = params.get("limit", None)
        filled = s.fillna(method="bfill", limit=limit)
        params["limit"] = limit
        return filled, params, True

    if method_label == LABEL_INTERP:
        # Prefer numeric; pandas may interpolate datetime under specific methods, but default to numeric
        if dtype_cat not in ("numeric", "datetime"):
            return s.copy(), params, False
        interp_method = params.get("interpolation_method", "linear")
        limit = params.get("limit", None)
        limit_direction = params.get("limit_direction", None)
        order = params.get("order", None)

        # Safely pass only provided kwargs
        kwargs: Dict[str, Any] = {"method": interp_method}
        if limit is not None:
            kwargs["limit"] = limit
        if limit_direction is not None:
            kwargs["limit_direction"] = limit_direction
        if order is not None:
            kwargs["order"] = order

        try:
            filled = s.interpolate(**kwargs)
        except Exception:
            # Interpolation failed for some reason
            return s.copy(), params, False

        params["interpolation_method"] = interp_method
        if limit is not None:
            params["limit"] = limit
        if limit_direction is not None:
            params["limit_direction"] = limit_direction
        if order is not None:
            params["order"] = order
        return filled, params, True

    if method_label == LABEL_DROP_ROWS:
        # Not implemented in the original script; preserve behavior (no deletion)
        # Mark as not applicable to avoid unintended changes.
        return s.copy(), params, False

    if method_label == LABEL_UNCHANGED:
        return s.copy(), params, True

    # Unknown label
    return s.copy(), params, False


def apply_with_fallback(
    s: pd.Series,
    chosen_method: str,
    dtype_cat: str,
    params: Optional[Dict[str, Any]],
    config: Optional[Dict[str, Any]],
) -> Tuple[pd.Series, str, bool, str, Dict[str, Any]]:
    """
    Применяет выбранный метод к серии с комплексными fallback-логиками по типам.
    Возвращает (заполненная_серия, финальный_метод, fallback_применен, описание_fallback, использованные_параметры).
    """
    params = dict(params or {})

    # Convenience for constants from config
    const_map: Dict[str, Any] = (config or {}).get("constants", {}) if config else {}
    default_const = (config or {}).get("default_constant", None) if config else None

    def try_once(m: str, p: Optional[Dict[str, Any]] = None) -> Tuple[pd.Series, str, bool, Dict[str, Any]]:
        filled, used, applicable = apply_method_once(s, m, p)
        return filled, m, applicable, used

    # Datetime-specific preference: ffill -> bfill -> unchanged
    if dtype_cat == "datetime":
        if chosen_method in (LABEL_FFILL, LABEL_BFILL, LABEL_INTERP):
            filled, m, ok, used = try_once(chosen_method, params)
            if ok:
                return filled, m, False, "", used
        # Fallback chain for datetime
        filled, m, ok, used = try_once(LABEL_FFILL, params)
        if ok and filled.isna().sum() < s.isna().sum():
            return filled, LABEL_FFILL, True, "Тип datetime → выбран ffill", used
        filled, m, ok, used = try_once(LABEL_BFILL, params)
        if ok and filled.isna().sum() < s.isna().sum():
            return filled, LABEL_BFILL, True, "Тип datetime: ffill не изменил → выбран bfill", used
        return s.copy(), LABEL_UNCHANGED, True, "Тип datetime: ffill/bfill неприменимы → Без изменений", {}

    # Boolean: prefer mode; else unchanged/constant if configured
    if dtype_cat == "boolean":
        if chosen_method != LABEL_MODE:
            # Fallback to mode for booleans
            filled, m, ok, used = try_once(LABEL_MODE, params)
            if ok:
                desc = f"Тип boolean: {chosen_method} неприменим → выбрана Мода"
                return filled, LABEL_MODE, True, desc, used
        # If chosen is mode, apply it
        filled, m, ok, used = try_once(LABEL_MODE, params)
        if ok:
            return filled, LABEL_MODE, False, "", used
        # No non-null values to infer mode
        const_val = const_map.get(s.name, default_const)
        if const_val is not None:
            filled, m, ok, used = try_once(LABEL_CONST, {"constant_value": const_val})
            if ok:
                return filled, LABEL_CONST, True, "Тип boolean: Мода отсутствует → используем Константу", used
        return s.copy(), LABEL_UNCHANGED, True, "Тип boolean: Мода/Константа недоступны → Без изменений", {}

    # String/categorical: prefer mode; then ffill; else unchanged
    if dtype_cat == "string":
        if chosen_method != LABEL_MODE:
            filled, m, ok, used = try_once(LABEL_MODE, params)
            if ok:
                desc = f"Строковой тип: {chosen_method} неприменим → выбрана Мода"
                return filled, LABEL_MODE, True, desc, used
        filled, m, ok, used = try_once(LABEL_MODE, params)
        if ok:
            return filled, LABEL_MODE, False, "", used
        # No mode available (all NaN) → ffill
        filled, m, ok, used = try_once(LABEL_FFILL, params)
        if ok and filled.isna().sum() < s.isna().sum():
            return filled, LABEL_FFILL, True, "Строковой тип: Мода отсутствует → выбран ffill", used
        return s.copy(), LABEL_UNCHANGED, True, "Строковой тип: Мода/ffill неприменимы → Без изменений", {}

    # Numeric: attempt chosen; then fallback chain Mode -> ffill -> unchanged
    filled, m, ok, used = try_once(chosen_method, params)
    if ok and filled.isna().sum() <= s.isna().sum():
        # Accept even if no reduction, as chosen method could be valid but not change values
        return filled, chosen_method, False, "", used

    # Fallback to mode
    filled, m, ok, used = try_once(LABEL_MODE, params)
    if ok:
        return filled, LABEL_MODE, True, f"Числовой тип: {chosen_method} неприменим → выбрана Мода", used

    # Then ffill
    filled, m, ok, used = try_once(LABEL_FFILL, params)
    if ok and filled.isna().sum() < s.isna().sum():
        return filled, LABEL_FFILL, True, f"Числовой тип: {chosen_method}/Мода неприменимы → выбран ffill", used

    # Else unchanged
    return s.copy(), LABEL_UNCHANGED, True, f"Числовой тип: {chosen_method}/Мода/ffill неприменимы → Без изменений", {}


def select_base_method_for_column(
    s: pd.Series,
    col_display_name: Any,
    total_rows: int,
    analysis_df: pd.DataFrame,
    config: Optional[Dict[str, Any]] = None,
) -> Tuple[str, Dict[str, Any]]:
    """
    Выбирает базовый метод по логике исходного скрипта (на основе column_analysis_results.xlsx)
    и с учётом возможной конфигурации.
    Возвращает (метод_меткой, параметры).
    """
    # Configuration can override per-column method
    if config and "methods" in config:
        by_col = config["methods"]
        if isinstance(by_col, dict):
            override = by_col.get(col_display_name)
            if override in ALLOWED_METHOD_LABELS:
                return override, {}

    # Original logic from fill_missing_values.py:
    # - Find corresponding row in analysis_df by column name (first column)
    # - If stats exist:
    #       if unique_count < 0.2 * total_rows and std_dev > 1 and precision > 0 -> use mean
    #   else: mode
    # - If no stats: mode
    stats = analysis_df[analysis_df.iloc[:, 0] == col_display_name]
    if not stats.empty:
        unique_count = stats.iloc[0, 1]
        std_dev = stats.iloc[0, 2]
        precision = stats.iloc[0, 3]
        if pd.notna(unique_count) and pd.notna(std_dev) and pd.notna(precision):
            try:
                if float(unique_count) < 0.2 * float(total_rows) and float(std_dev) > 1 and float(precision) > 0:
                    return LABEL_MEAN, {}
            except Exception:
                # If conversion fails, fallback to mode
                pass
        return LABEL_MODE, {}
    else:
        return LABEL_MODE, {}


# ---------------------------- Core API ----------------------------

def apply_filling(df: pd.DataFrame, config: Optional[Dict[str, Any]] = None) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    """
    Применяет стратегии заполнения к копии df, возвращает (filled_df, log_df, summary_dict).
    - filled_df: заполненный DataFrame (порядок колонок сохранён)
    - log_df: таблица лога по колонкам
    - summary_dict: агрегированные показатели для листа "Сводка"
    """
    df_filled = df.copy(deep=True)

    # Attempt to load external analysis to preserve original behavior
    # Note: The original script reads 'column_analysis_results.xlsx' directly; keep this default.
    try:
        analysis_df = pd.read_excel('column_analysis_results.xlsx')
    except Exception:
        # If file is missing, construct an empty DataFrame so selection defaults to "Мода"
        analysis_df = pd.DataFrame(columns=["col", "unique", "std", "precision"])

    total_rows = len(df_filled)

    log_rows: List[Dict[str, Any]] = []
    methods_counter: Counter = Counter()

    # Optional default params from config
    ffill_limit = (config or {}).get("ffill_limit")
    bfill_limit = (config or {}).get("bfill_limit")
    interp_defaults: Dict[str, Any] = (config or {}).get("interpolate", {}) if config else {}

    # Iterate preserving column order; handle MultiIndex by using the tuple directly for series
    for col in df_filled.columns:
        s = df_filled[col]
        # Display name is first level for compatibility with original logic
        col_display_name = col[0] if isinstance(col, tuple) else col
        dtype_cat = get_dtype_category(s)
        missing_before = int(s.isna().sum())

        start_t = time.perf_counter()

        if missing_before == 0:
            # No changes needed
            filled_series = s.copy()
            final_method = LABEL_UNCHANGED
            fallback_applied = False
            fallback_desc = ""
            used_params: Dict[str, Any] = {}
        else:
            # Select base method as original script
            base_method, base_params = select_base_method_for_column(
                s, col_display_name, total_rows, analysis_df, config
            )

            # Provide sensible defaults for params depending on method
            params: Dict[str, Any] = dict(base_params)
            if base_method == LABEL_FFILL and ffill_limit is not None:
                params["limit"] = ffill_limit
            if base_method == LABEL_BFILL and bfill_limit is not None:
                params["limit"] = bfill_limit
            if base_method == LABEL_INTERP:
                params.update(interp_defaults)

            # Apply with dtype-aware fallback
            filled_series, final_method, fallback_applied, fallback_desc, used_params = apply_with_fallback(
                s, base_method, dtype_cat, params, config
            )

        # Assign back without mutating original series
        df_filled[col] = filled_series

        elapsed_ms = (time.perf_counter() - start_t) * 1000.0
        missing_after = int(filled_series.isna().sum())

        # Update method distribution
        methods_counter[final_method] += 1

        # Build log row
        # Combine parameters into readable string
        if used_params:
            # Keep deterministic order
            param_items = []
            for k in sorted(used_params.keys()):
                param_items.append(f"{k}={used_params[k]}")
            params_str = ", ".join(param_items)
        else:
            params_str = ""

        log_rows.append(
            {
                "Название колонки": str(col_display_name),
                "Метод заполнения": final_method,
                "Параметры метода": params_str,
                "Пропусков до": missing_before,
                "Пропусков после": missing_after,
                "Тип данных столбца": dtype_cat,
                "Fallback применен": bool(fallback_applied),
                "Описание fallback": fallback_desc,
                "Время обработки (мс)": round(elapsed_ms, 3),
            }
        )

    # Build summary
    total_missing_before = int(df.isna().sum().sum())
    total_missing_after = int(df_filled.isna().sum().sum())
    total_time_ms = round(sum(r["Время обработки (мс)"] for r in log_rows), 3)

    summary_dict: Dict[str, Any] = {
        "Общее число пропусков до": total_missing_before,
        "Общее число пропусков после": total_missing_after,
        "Количество обработанных столбцов": int(len(df_filled.columns)),
        "Распределение применённых методов": dict(methods_counter),
        "Общее время выполнения (мс)": total_time_ms,
    }

    log_df, summary_df = build_filling_log(log_rows, summary_dict)
    return df_filled, log_df, summary_dict


def build_filling_log(log_rows: List[Dict[str, Any]], summary: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Строит таблицы лога и сводки.
    Возвращает (log_df, summary_df).
    """
    log_df = pd.DataFrame(log_rows)

    # Build "Сводка" sheet as a two-column table "Показатель"/"Значение"
    # plus collapsed distribution block.
    summary_rows: List[Dict[str, Any]] = [
        {"Показатель": "Общее число пропусков до", "Значение": summary.get("Общее число пропусков до", 0)},
        {"Показатель": "Общее число пропусков после", "Значение": summary.get("Общее число пропусков после", 0)},
        {"Показатель": "Количество обработанных столбцов", "Значение": summary.get("Количество обработанных столбцов", 0)},
        {"Показатель": "Общее время выполнения (мс)", "Значение": summary.get("Общее время выполнения (мс)", 0)},
        {"Показатель": "", "Значение": ""},  # separator
        {"Показатель": "Распределение применённых методов:", "Значение": ""},
    ]

    methods_dist = summary.get("Распределение применённых методов", {})
    for method_label in sorted(methods_dist.keys()):
        summary_rows.append(
            {
                "Показатель": f"  - {method_label}",
                "Значение": methods_dist[method_label],
            }
        )

    summary_df = pd.DataFrame(summary_rows)
    return log_df, summary_df


# ---------------------------- Entry point ----------------------------

def main() -> None:
    """
    Основная точка входа.
    - Читает входные файлы как в исходном скрипте:
        * 'result_no_text_filled_more_than_80.xlsx' с multi-index заголовками (первые 3 строки)
        * 'column_analysis_results.xlsx' (используется внутри apply_filling)
    - Применяет заполнение, сохраняет результат и лог.
    """
    # Preserve original paths and behavior
    input_path = 'result_no_text_filled_more_than_80.xlsx'
    output_path = 'result_no_text_filled_more_than_80_filled.xlsx'

    # Load main DataFrame with MultiIndex headers from first 3 rows
    df = pd.read_excel(input_path, header=[0, 1, 2])

    # Apply filling with default behavior (no external config)
    filled_df, log_df, summary = apply_filling(df, config=None)

    # Save main result preserving first 3 header rows via openpyxl, as in original script
    try:
        wb = load_workbook(input_path)
        ws = wb.active

        # Write filled data starting from row 4 (1-based indexing): original data starts at row 4
        # Note: Do not touch the first 3 header rows.
        n_rows = len(filled_df)
        n_cols = len(filled_df.columns)

        for r in range(n_rows):
            for c in range(n_cols):
                ws.cell(row=r + 4, column=c + 1, value=filled_df.iloc[r, c])

        # Save to fixed output path (preserving behavior)
        wb.save(output_path)
    except Exception as e:
        print(f"Ошибка при сохранении основного результата: {e}")

    # Compose log path next to main output: <base>_filling_methods_log.xlsx
    if output_path.lower().endswith(".xlsx"):
        log_path = output_path[:-5] + "_filling_methods_log.xlsx"
    else:
        log_path = output_path + "_filling_methods_log.xlsx"

    # Build both log tables
    log_table_df, summary_df = build_filling_log([dict(row) for _, row in log_df.iterrows()], summary)

    # Save the log workbook with two sheets ("Лог" and "Сводка") using openpyxl
    try:
        with pd.ExcelWriter(log_path, engine="openpyxl") as writer:
            log_table_df.to_excel(writer, index=False, sheet_name="Лог")
            summary_df.to_excel(writer, index=False, sheet_name="Сводка")
    except Exception as e:
        print(f"Ошибка при сохранении лога: {e}")


if __name__ == "__main__":
    # Protective entry point, preserving backward-compatible behavior (no CLI args required).
    main()